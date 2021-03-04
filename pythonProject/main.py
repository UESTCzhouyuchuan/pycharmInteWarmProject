from openpyxl import *

def main():
    wb = load_workbook(filename='四川省地图填色(1).xlsx')
    ws = wb[wb.sheetnames[1]]
    for row in ws.iter_rows():
        for cell in row:
            print(cell.value)
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
