import json
import re

import openpyxl

from printState import *


def writeToFile(filename, data, encoding="UTF-8"):
    try:
        file = open(filename, mode='w+', encoding=encoding)
    except Exception as e:
        printError(f"can't open {filename}, error: {e}")
    else:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.close()
        printSuccess('数据写入文件' + filename)


def readFromFile(filename, encoding="UTF-8"):
    try:
        file = open(filename, mode='r', encoding=encoding)
    except Exception as e:
        printError(f"can't open {filename}, error: {e}")
    else:
        data = json.load(file)
        file.close()
        printSuccess('读入数据文件' + filename)
        return data


def re_strip(s, t=r'\s'):
    t_format = rf'^{t}*|{t}*$'
    s_re = re.compile(t_format)
    s = s_re.sub('', s)
    return s


def ILLEGAL_CHARACTERS(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
    text = text.replace('<![CDATA[<div>', '')
    text = text.replace('\xa0', ' ')
    text = text.strip()
    return text


def writeToXlsx(filename, data, rows):
    try:
        wb = openpyxl.Workbook()
    except Exception as e:
        printError(f"创建{filename},error: {e}")
    ws = wb.active
    ws.delete_rows(1, ws.max_row)

    headers = rows[:]
    headers.insert(0, 'number')
    for i in range(len(headers)):
        ws.cell(1, i + 1).value = headers[i]
        ws.cell(1, i + 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    row = 0
    for i in range(len(data)):
        item = data[i]
        if item is not None:
            try:
                ws.cell(row + 2, 1).value = row + 1
                ws.cell(row + 2, 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                for j in range(1, len(headers)):
                    ws.cell(row + 2, j + 1).value = str(item[headers[j]])
                row += 1
            except Exception as e:
                printError(f"写入第{row}行数据到{filename},{item}, error: {e}")
    try:
        wb.save(filename)
    except Exception as e:
        printError(f"保存文件{filename}, error: {e}")
    else:
        printSuccess(f"写入数据到{filename}")
    wb.close()
